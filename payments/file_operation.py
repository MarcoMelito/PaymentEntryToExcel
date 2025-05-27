import shutil
import os
from openpyxl import load_workbook


# Get the data from ui, search or create the daily folder,
# save the daily file in the directory and update the daily file and all_payments file
class Operation:
    def __init__(self):
        self.payments_file = None

    # Open File excel
    def open_file(self, input_date) -> None:
        # Create the daily file path
        path_file = f"{input_date}./PAYMENTS_FILE.xlsx"
        # Check if the path already exist if not create the folder and copy the blank file in the new folder
        if not os.path.exists(path_file):
            os.makedirs(f"{input_date}./")
            shutil.copyfile('PAYMENTS_FILE.xlsx', f"{path_file}")

        # This part of script scans the daily folders and automatically moves them to their respective monthly folders,
        # creating them if necessary.
        # It iterates over all months of the year (January to December)
        for month in range(1, 13):
            # iterates over all days of the month (1 through 31)
            for day in range(1, 32):
                # Construction of daily folder name (move_path).
                # This logic adds a leading zero to day and month if they are less than 10
                if day < 10 and month < 10:
                    # Example: 01_01_2025
                    move_path = f"0{day}_0{month}{input_date[5:]}"
                elif day < 10 and month > 9:
                    # Example: 01_10_2025
                    move_path = f"0{day}_{month}{input_date[5:]}"
                elif day > 9 and month < 10:
                    # Example: 10_01_2025
                    move_path = f"{day}_0{month}{input_date[5:]}"
                else:
                    # Example: 10_10_2025
                    move_path = f"{day}_{month}{input_date[5:]}"
                # Construction of the target monthly folder name (month_path)
                month_path = f"{move_path[3:]}"
                # Check if the daily folder exists in the current directory
                if os.path.exists(move_path):
                    # Condition to avoid moving the current folder
                    if move_path != path_file.split("./")[0]:
                        # Checks if the monthly target folder exists.
                        if not os.path.exists(month_path):
                            # If it doesn't exist, it creates it
                            os.makedirs(f"{month_path}./")
                        # Move daily folder to monthly folder
                        shutil.move(f'./{move_path}', f"{month_path}./{move_path}")

        # Load the Excel file
        self.payments_file = load_workbook(filename=f"{input_date}./PAYMENTS_FILE.xlsx")

    # Update the daily Excel file with data got from UI
    def update_file(self, entry) -> None:
        # Date extraction to get the correct path
        date_input = entry[0].replace("/", "_")
        self.open_file(input_date=date_input)
        # Load the daily file
        file_to_update = load_workbook(f"{date_input}./PAYMENTS_FILE.xlsx")
        # Get the sheet
        sheet = file_to_update["Foglio1"]
        # Call the static method to enter the 'entry' data in the first empty row of the sheet
        self.insert_data(entry, sheet)
        # Save the file
        file_to_update.save(f"{date_input}./PAYMENTS_FILE.xlsx")
        # Call the function for save all the payments
        self.save_all_payments(entry)
        return True

    # Update the file all_payments
    def save_all_payments(self, entry) -> None:
        # Load the file all payments
        file_to_update = load_workbook("all_payments./PAYMENTS_FILE.xlsx")
        # Get the sheet
        sheet = file_to_update["Foglio1"]
        # Call the static method to enter the 'entry' data in the first empty row of the sheet
        self.insert_data(entry, sheet)
        # Save the file
        file_to_update.save("all_payments./PAYMENTS_FILE.xlsx")

    # Method that iterate on column and row of the sheet and enters the data into the file
    @staticmethod
    def insert_data(entry, sheet):
        # Get the number of row
        max_row = sheet.max_row
        # Iterate on column
        for i in range(1, 8):
            # Iterate on row
            for row in range(2, max_row):
                # Check if row is empty, if true enters the data
                if sheet.cell(row, i).value is None:
                    sheet.cell(row=row, column=i, value=entry[i - 1])
                    break
